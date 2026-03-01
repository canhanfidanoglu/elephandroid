from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import openpyxl

from .models import ColumnMapping, ParsedTask

PRIORITY_MAP = {
    "critical": 1,
    "urgent": 1,
    "high": 3,
    "medium": 5,
    "normal": 5,
    "low": 9,
}


def _cell_value(ws, row, col_letter):
    cell = ws[f"{col_letter}{row}"]
    val = cell.value
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        return val if val else None
    return val


def _parse_priority(raw):
    if raw is None:
        return None
    text = str(raw).strip().lower()
    if not text:
        return None
    if text in PRIORITY_MAP:
        return PRIORITY_MAP[text]
    try:
        num = int(text)
        if num in PRIORITY_MAP.values():
            return num
    except (ValueError, TypeError):
        pass
    return None


def _parse_date(raw):
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _to_str_or_none(val):
    if val is None:
        return None
    return str(val)


def _parse_sheet(ws, mapping: ColumnMapping) -> list[ParsedTask]:
    tasks = []
    current_task = None

    for row in range(mapping.start_row, ws.max_row + 1):
        ticket_id = _cell_value(ws, row, mapping.ticket_id)
        checklist_item = _cell_value(ws, row, mapping.checklist)

        if ticket_id:
            # This is a main task row
            if current_task:
                tasks.append(current_task)

            title = _cell_value(ws, row, mapping.title)
            if not title:
                continue

            current_task = ParsedTask(
                ticket_id=str(ticket_id),
                title=str(title),
                epic=_to_str_or_none(_cell_value(ws, row, mapping.epic)),
                description=_to_str_or_none(_cell_value(ws, row, mapping.notes)),
                bucket_name=_to_str_or_none(_cell_value(ws, row, mapping.bucket)),
                priority=_parse_priority(_cell_value(ws, row, mapping.priority)),
                start_date=_parse_date(_cell_value(ws, row, mapping.start_date)),
                due_date=_parse_date(_cell_value(ws, row, mapping.due_date)),
                assignee=_to_str_or_none(_cell_value(ws, row, mapping.assignee)),
                checklist_items=[],
            )
        elif checklist_item and current_task:
            # This is a checklist sub-row for the current task
            current_task.checklist_items.append(str(checklist_item))

    # Don't forget the last task
    if current_task:
        tasks.append(current_task)

    return tasks


def parse_excel(
    file_path_or_bytes, mapping: ColumnMapping = ColumnMapping()
) -> list[ParsedTask]:
    if isinstance(file_path_or_bytes, (str, Path)):
        wb = openpyxl.load_workbook(
            file_path_or_bytes, read_only=True, data_only=True
        )
    elif isinstance(file_path_or_bytes, (bytes, BytesIO)):
        buf = (
            BytesIO(file_path_or_bytes)
            if isinstance(file_path_or_bytes, bytes)
            else file_path_or_bytes
        )
        wb = openpyxl.load_workbook(buf, data_only=True)
    else:
        raise TypeError(
            f"Expected file path, bytes, or BytesIO, got {type(file_path_or_bytes)}"
        )

    try:
        all_tasks = []

        if mapping.sheet_name:
            sheets = [wb[mapping.sheet_name]]
        else:
            # Parse all sheets (skip non-data sheets like "Timeline Overview")
            sheets = [
                ws
                for ws in wb.worksheets
                if ws.max_row and ws.max_row > 1
            ]

        for ws in sheets:
            # Check if sheet has the expected header structure
            header_check = _cell_value(ws, 1, mapping.ticket_id)
            if header_check and "ticket" in str(header_check).lower():
                sheet_tasks = _parse_sheet(ws, mapping)
                # If sheet_as_bucket mode: override bucket_name with sheet title
                if mapping.sheet_as_bucket and ws.title:
                    for t in sheet_tasks:
                        if not t.bucket_name:
                            t.bucket_name = ws.title
                all_tasks.extend(sheet_tasks)

        return all_tasks
    finally:
        wb.close()
